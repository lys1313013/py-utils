import os
from openpyxl import load_workbook, Workbook

def split_excel(file_path, rows_per_file):
    """
    将 Excel 文件按指定行数拆分为多个子文件。
    :param file_path: 原始 Excel 文件路径
    :param rows_per_file: 每个子文件的数据行数（不含表头）
    """
    if not os.path.exists(file_path):
        print(f"错误: 找不到文件 {file_path}")
        return

    # 准备输出目录
    file_dir = os.path.dirname(os.path.abspath(file_path))
    file_name = os.path.basename(file_path)
    name_wo_ext, ext = os.path.splitext(file_name)
    output_dir = os.path.join(file_dir, f"{name_wo_ext}_split")
    os.makedirs(output_dir, exist_ok=True)

    # 加载工作簿和表头
    wb = load_workbook(file_path, data_only=True, read_only=True)
    ws = wb.active
    
    # 获取表头
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        print("文件为空")
        return

    part = 1
    current_row_count = 0
    new_wb = None

    for row in rows_iter:
        if current_row_count == 0:
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.append(header)
        
        new_ws.append(row)
        current_row_count += 1
        
        if current_row_count >= rows_per_file:
            save_path = os.path.join(output_dir, f"{name_wo_ext}_part{part}{ext}")
            new_wb.save(save_path)
            print(f"已保存: {save_path}")
            part += 1
            current_row_count = 0
            new_wb = None

    # 保存最后一组数据
    if new_wb:
        save_path = os.path.join(output_dir, f"{name_wo_ext}_part{part}{ext}")
        new_wb.save(save_path)
        print(f"已保存: {save_path}")

    wb.close()
    print(f"\n所有文件已保存至: {output_dir}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 3:
        print("用法: python split_excel.py <文件路径> <每页行数>")
    else:
        split_excel(sys.argv[1], int(sys.argv[2]))
